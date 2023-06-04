from datetime import datetime
from dateutil.relativedelta import relativedelta
import collections
import json
import os
import openpyxl
class Report:
    def __init__(self,tree):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "worklogs"
        i=2
        header={}
        rows=tree.getrows()
        for id in rows:
            row=rows[id]
           
            if(row.data != None):
                j=1
                #print(row.data['key'],row.worklogs)
                
                
                sheet.cell(i,j).value=row.data['key']
                header['Jira']='Jira'
                j=j+1
                
                sheet.cell(i,j).value=row.data['summary']
                header['Summary']='Summary'
                j=j+1
                
                #sheet.cell(i,j).value=row.data['summary']
                header['Resource']='Resource'
                resource_col=j
                j=j+1
                
                #if row.data["key"]=='PSP-9265':
                #    print(row.worklogs)
                    
                for dd in row.worklogs:
                    header[dd]=dd
                    if row.worklogs[dd]['total']!=0:
                        sheet.cell(i,3).value='Total'
                        sheet.cell(i,j).value=f"{row.worklogs[dd]['total']} Hrs"
                        #if row.data["key"]=='PSP-9265':
                        #    print(i)
                        #for m in row.worklogs[dd]:
                        #    if m=='total':
                        #        pass
                        #    else:
                        #       sheet.cell(i,j).value=f"{row.worklogs[dd][m]} Hrs"
                        #       i=i+1
                        
                    j=j+1
                i=i+1
                row_rem=i
                max=row_rem
                j=3
                for dd in row.worklogs:
                    i=row_rem
                    j=j+1
                    #print(row.worklogs)
                    for m in row.worklogs[dd]:
                        if m=='total':
                            #if row.data["key"]=='PSP-9265':
                            #    print(i,"-->"+m)
                            pass
                        else:
                            #if row.data["key"]=='PSP-9265':
                            #    print(i,m)
                            sheet.cell(i,3).value=m
                            sheet.cell(i,j).value=f"{row.worklogs[dd][m]} Hrs"
                            i=i+1    
                            if i>max:
                                max=i
                            #if row.data["key"]=='PSP-9265':
                            #    print(i,m)
                i=max          
                #if row.data["key"]=='PSP-9265':
                #    wb.save(f'/projects/{str(tree.id)}/report.xlsx') 
                #    exit()
                
        j=1
        for field in header:
            sheet.cell(1,j).value=field
            j=j+1
       
        i=2
        header={}
        sheet=wb.create_sheet('schedule')
        for id in rows:
            row=rows[id]
           
            if(row.data != None):
                j=1
                #print(row.data['key'],row.schedule)
                sheet.cell(i,j).value=row.data['key']
                header['Jira']='Jira'
                j=j+1
                
                sheet.cell(i,j).value=row.data['summary']
                header['Summary']='Summary'
                j=j+1
                
                for dd in row.schedule:
                    header[dd]=dd
                    if row.schedule[dd]!=0:
                        sheet.cell(i,j).value=f"{row.schedule[dd]} Hrs"
                    j=j+1
                i=i+1
        j=1
        for field in header:
            sheet.cell(1,j).value=field
            j=j+1 
        
        wb.save(f'/projects/{str(tree.id)}/report.xlsx') 
        print("Report generated")
